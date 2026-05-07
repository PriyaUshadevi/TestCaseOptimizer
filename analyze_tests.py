"""
Test Suite Duplicate Detector

Reads an Excel test-case file, compares test cases for duplicate or near-duplicate
coverage, and writes a recruiter/client-friendly Excel report.
"""

from __future__ import annotations

import argparse
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_FILE = "test_cases.xlsx"
EXACT_THRESHOLD = 0.92
SIMILAR_THRESHOLD = 0.70

RED = "FFD7D7"
ORANGE = "FFE8CC"
GREEN = "D7F5E3"
HEADER = "1F4E79"
WHITE = "FFFFFF"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze an Excel test suite and flag duplicate test cases."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        default=DEFAULT_INPUT_FILE,
        help=f"Excel file to analyze. Defaults to {DEFAULT_INPUT_FILE}.",
    )
    parser.add_argument(
        "--output",
        help="Optional output Excel filename. Defaults to <input>_DUPLICATES_FLAGGED.xlsx.",
    )
    parser.add_argument(
        "--exact-threshold",
        type=float,
        default=EXACT_THRESHOLD,
        help="Similarity score for exact duplicates. Default: 0.92.",
    )
    parser.add_argument(
        "--similar-threshold",
        type=float,
        default=SIMILAR_THRESHOLD,
        help="Similarity score for near duplicates. Default: 0.70.",
    )
    return parser.parse_args()


def similarity(a: str, b: str) -> float:
    """Return a 0-1 similarity score between two strings."""
    if not a or not b:
        return 0.0
    left = a.lower().strip()
    right = b.lower().strip()
    if left == right:
        return 1.0

    left_tokens = set(left.split())
    right_tokens = set(right.split())
    token_overlap = len(left_tokens & right_tokens) / max(len(left_tokens | right_tokens), 1)
    sequence_score = SequenceMatcher(None, left, right).ratio()
    return (sequence_score * 0.7) + (token_overlap * 0.3)


def clean(value: object) -> str:
    return str(value).strip() if pd.notna(value) else ""


def find_column(columns: list[str], candidates: list[str], contains: list[str] | None = None) -> str | None:
    normalized = {column.lower().strip(): column for column in columns}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]

    contains = contains or []
    for column in columns:
        lower = column.lower().strip()
        if any(text in lower for text in contains):
            return column
    return None


def map_columns(df: pd.DataFrame) -> dict[str, str | None]:
    columns = list(df.columns)
    return {
        "id": find_column(
            columns,
            ["test case id", "testcaseid", "tc id", "id"],
            ["test case id"],
        ),
        "summary": find_column(
            columns,
            ["summary", "test case name", "test name", "title", "name"],
            ["summary", "test case name", "test name"],
        ),
        "area": find_column(columns, ["test case area", "area", "module", "feature"], ["area"]),
        "description": find_column(columns, ["description", "test description"], ["description"]),
    }


def output_filename(input_file: Path, explicit_output: str | None) -> Path:
    if explicit_output:
        return Path(explicit_output)
    return input_file.with_name(f"{input_file.stem}_DUPLICATES_FLAGGED.xlsx")


def build_cases(df: pd.DataFrame, col_map: dict[str, str | None]) -> list[dict[str, str]]:
    id_col = col_map["id"]
    summary_col = col_map["summary"]
    area_col = col_map.get("area")
    desc_col = col_map.get("description")

    if not id_col or not summary_col:
        missing = []
        if not id_col:
            missing.append("Test Case ID")
        if not summary_col:
            missing.append("Summary or Test Case Name")
        raise ValueError(
            "Missing required column(s): "
            + ", ".join(missing)
            + f". Columns found: {list(df.columns)}"
        )

    grouped = df.groupby(id_col, sort=False).first().reset_index()
    cases = []
    for _, row in grouped.iterrows():
        summary = clean(row[summary_col])
        description = clean(row[desc_col]) if desc_col else ""
        cases.append(
            {
                "id": clean(row[id_col]),
                "summary": summary,
                "area": clean(row[area_col]) if area_col else "",
                "description": description,
                "combined": f"{summary} {description}".strip(),
            }
        )
    return cases


def analyze_cases(
    cases: list[dict[str, str]],
    exact_threshold: float,
    similar_threshold: float,
) -> tuple[list[tuple[str, str, str, str, float, str]], dict[str, str]]:
    duplicate_pairs = []
    flagged_ids: dict[str, str] = {}

    for i, case_a in enumerate(cases):
        for case_b in cases[i + 1 :]:
            score = similarity(case_a["combined"], case_b["combined"])
            if score >= exact_threshold:
                duplicate_type = "EXACT DUPLICATE"
            elif score >= similar_threshold:
                duplicate_type = "NEAR DUPLICATE"
            else:
                continue

            for case_id in [case_a["id"], case_b["id"]]:
                if duplicate_type == "EXACT DUPLICATE" or case_id not in flagged_ids:
                    flagged_ids[case_id] = duplicate_type

            duplicate_pairs.append(
                (
                    case_a["id"],
                    case_a["summary"],
                    case_b["id"],
                    case_b["summary"],
                    round(score * 100, 1),
                    duplicate_type,
                )
            )

    duplicate_pairs.sort(key=lambda pair: -pair[4])
    return duplicate_pairs, flagged_ids


def style_sheet(ws, flag_col_name: str | None = None, type_col_name: str | None = None) -> None:
    header_fill = PatternFill("solid", fgColor=HEADER)
    header_font = Font(bold=True, color=WHITE, size=10)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 30

    flag_idx = None
    type_idx = None
    for index, cell in enumerate(ws[1], 1):
        if cell.value == flag_col_name:
            flag_idx = index
        if cell.value == type_col_name:
            type_idx = index

    fill_map = {
        "EXACT DUPLICATE": PatternFill("solid", fgColor=RED),
        "NEAR DUPLICATE": PatternFill("solid", fgColor=ORANGE),
        "UNIQUE - KEEP": PatternFill("solid", fgColor=GREEN),
    }

    for row in ws.iter_rows(min_row=2):
        status = None
        if flag_idx and row[flag_idx - 1].value:
            status = str(row[flag_idx - 1].value)
        if type_idx and row[type_idx - 1].value:
            status = str(row[type_idx - 1].value)

        row_fill = fill_map.get(status)
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_fill:
                cell.fill = row_fill

    for col in ws.iter_cols():
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_summary_sheet(
    workbook_path: Path,
    input_file: Path,
    total_cases: int,
    exact_count: int,
    near_count: int,
    flagged_count: int,
) -> None:
    wb = load_workbook(workbook_path)
    style_sheet(wb["All Tests - Flagged"], flag_col_name="DUPLICATE STATUS")
    style_sheet(wb["Duplicate Pairs Report"], type_col_name="Type")

    if "SUMMARY" in wb.sheetnames:
        del wb["SUMMARY"]
    ws = wb.create_sheet("SUMMARY", 0)

    estimated_remove = exact_count + near_count // 2
    summary_data = [
        ["TEST SUITE DUPLICATE ANALYSIS REPORT", ""],
        ["", ""],
        ["INPUT FILE", str(input_file)],
        ["TOTAL UNIQUE TEST CASES", total_cases],
        ["EXACT DUPLICATE PAIRS", exact_count],
        ["NEAR DUPLICATE PAIRS", near_count],
        ["TEST CASES FLAGGED", flagged_count],
        ["ESTIMATED SAFE TO REMOVE", estimated_remove],
        ["RECOMMENDED FINAL COUNT", max(total_cases - estimated_remove, 0)],
        ["", ""],
        ["COLOUR GUIDE", ""],
        ["RED", "Exact duplicate - safe candidate to remove one after review"],
        ["ORANGE", "Near duplicate - review before merging or deleting"],
        ["GREEN", "Unique - keep"],
        ["", ""],
        ["NEXT STEPS", ""],
        ["1", "Review the Duplicate Pairs Report tab, highest Match % first"],
        ["2", "Choose the best version of each duplicate test case to keep"],
        ["3", "Update the master test suite or Jira/Xray import file"],
    ]

    for row_index, row_data in enumerate(summary_data, 1):
        for col_index, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="center")
            if row_index == 1:
                cell.font = Font(bold=True, size=14, color=HEADER)
            elif value in {
                "INPUT FILE",
                "TOTAL UNIQUE TEST CASES",
                "EXACT DUPLICATE PAIRS",
                "NEAR DUPLICATE PAIRS",
                "TEST CASES FLAGGED",
                "ESTIMATED SAFE TO REMOVE",
                "RECOMMENDED FINAL COUNT",
                "COLOUR GUIDE",
                "NEXT STEPS",
            }:
                cell.font = Font(bold=True, size=10)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70
    wb.save(workbook_path)


def write_report(
    df: pd.DataFrame,
    output_file: Path,
    id_col: str,
    duplicate_pairs: list[tuple[str, str, str, str, float, str]],
    flagged_ids: dict[str, str],
) -> None:
    report_df = df.copy()
    report_df["DUPLICATE STATUS"] = report_df[id_col].map(
        lambda value: flagged_ids.get(clean(value), "UNIQUE - KEEP")
    )

    pairs_df = pd.DataFrame(
        duplicate_pairs,
        columns=[
            "Test ID (A)",
            "Summary (A)",
            "Test ID (B)",
            "Summary (B)",
            "Match %",
            "Type",
        ],
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name="All Tests - Flagged", index=False)
        pairs_df.to_excel(writer, sheet_name="Duplicate Pairs Report", index=False)


def main() -> int:
    args = parse_args()
    input_file = Path(args.input_file)

    print("\n" + "=" * 60)
    print("  TEST SUITE DUPLICATE DETECTOR")
    print("=" * 60)

    if not input_file.exists():
        print(f"\nERROR: Cannot find file '{input_file}'.")
        print("Place the Excel file in this folder or pass the full path to the file.")
        return 1

    print(f"\nLoading: {input_file}")
    df = pd.read_excel(input_file, dtype=str)
    df.columns = [str(column).strip() for column in df.columns]
    print(f"Rows loaded: {len(df)}")
    print(f"Columns found: {list(df.columns)}")

    try:
        col_map = map_columns(df)
        cases = build_cases(df, col_map)
    except ValueError as exc:
        print(f"\nERROR: {exc}")
        return 1

    print("\nMapped columns:")
    for key, value in col_map.items():
        if value:
            print(f"  {key:12} -> {value}")

    print(f"\nUnique test cases found: {len(cases)}")
    print("Comparing test cases...")
    duplicate_pairs, flagged_ids = analyze_cases(
        cases,
        args.exact_threshold,
        args.similar_threshold,
    )

    exact_count = sum(1 for pair in duplicate_pairs if pair[5] == "EXACT DUPLICATE")
    near_count = sum(1 for pair in duplicate_pairs if pair[5] == "NEAR DUPLICATE")
    output_file = output_filename(input_file, args.output)

    write_report(df, output_file, col_map["id"], duplicate_pairs, flagged_ids)
    add_summary_sheet(
        output_file,
        input_file,
        len(cases),
        exact_count,
        near_count,
        len(flagged_ids),
    )

    print("\nAnalysis complete.")
    print(f"Total unique test cases : {len(cases)}")
    print(f"Exact duplicate pairs   : {exact_count}")
    print(f"Near duplicate pairs    : {near_count}")
    print(f"Test cases flagged      : {len(flagged_ids)}")
    print(f"Output file saved       : {output_file}")
    print("\nOpen the output Excel file and start with the SUMMARY tab.")
    print("=" * 60 + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
