from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from analyze_tests import (
    analyze_cases,
    build_cases,
    map_columns,
    output_filename,
    similarity,
    validate_thresholds,
    write_report,
    add_summary_sheet,
)


def test_similarity_scores_exact_and_unrelated_text():
    assert similarity("Login Test", "login test") == 1.0
    assert similarity("Create supplier invoice", "Reset customer password") < 0.7


def test_threshold_validation_rejects_invalid_order():
    try:
        validate_thresholds(0.5, 0.8)
    except ValueError as exc:
        assert "--similar-threshold" in str(exc)
    else:
        raise AssertionError("Expected invalid threshold order to raise ValueError")


def test_build_cases_maps_common_columns_and_groups_by_id():
    df = pd.DataFrame(
        [
            {"TC ID": "TC-1", "Summary": "Login works", "Description": "Valid user"},
            {"TC ID": "TC-1", "Summary": "Ignored duplicate row", "Description": "Step row"},
            {"TC ID": "TC-2", "Summary": "Logout works", "Description": "End session"},
        ]
    )

    col_map = map_columns(df)
    cases = build_cases(df, col_map)

    assert col_map["id"] == "TC ID"
    assert col_map["summary"] == "Summary"
    assert len(cases) == 2
    assert cases[0]["id"] == "TC-1"


def test_analyze_cases_flags_exact_duplicates():
    cases = [
        {"id": "TC-1", "summary": "Login Test", "combined": "Login Test"},
        {"id": "TC-2", "summary": "Login Test", "combined": "Login Test"},
        {"id": "TC-3", "summary": "Logout Test", "combined": "Logout Test"},
    ]

    pairs, flagged = analyze_cases(cases, exact_threshold=0.92, similar_threshold=0.7)

    assert len(pairs) == 1
    assert pairs[0][5] == "EXACT DUPLICATE"
    assert flagged == {"TC-1": "EXACT DUPLICATE", "TC-2": "EXACT DUPLICATE"}


def test_report_writer_creates_expected_workbook(tmp_path):
    df = pd.DataFrame(
        [
            {"Test Case ID": "TC-1", "Test Case Name": "Login Test"},
            {"Test Case ID": "TC-2", "Test Case Name": "Login Test"},
            {"Test Case ID": "TC-3", "Test Case Name": "Logout Test"},
        ]
    )
    output_file = tmp_path / "reports" / "flagged.xlsx"
    duplicate_pairs = [
        ("TC-1", "Login Test", "TC-2", "Login Test", 100.0, "EXACT DUPLICATE")
    ]
    flagged = {"TC-1": "EXACT DUPLICATE", "TC-2": "EXACT DUPLICATE"}

    write_report(df, output_file, "Test Case ID", duplicate_pairs, flagged)
    add_summary_sheet(output_file, Path("test_cases.xlsx"), 3, 1, 0, 2)

    workbook = load_workbook(output_file)
    assert workbook.sheetnames == ["SUMMARY", "All Tests - Flagged", "Duplicate Pairs Report"]
    assert workbook["All Tests - Flagged"]["C2"].value == "EXACT DUPLICATE"
    assert workbook["All Tests - Flagged"]["C4"].value == "UNIQUE - KEEP"


def test_output_filename_uses_input_stem_when_no_explicit_name():
    assert output_filename(Path("sample.xlsx"), None) == Path("sample_DUPLICATES_FLAGGED.xlsx")
    assert output_filename(Path("sample.xlsx"), "custom.xlsx") == Path("custom.xlsx")
